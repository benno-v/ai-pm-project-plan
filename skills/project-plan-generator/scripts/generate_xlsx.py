#!/usr/bin/env python3
"""
Generate the Excel workbook for a project plan.

Sheets produced:
    - Overview     : project meta, executive snapshot, status summary, optional logo.
    - Gantt        : phase-grouped, weekly-resolution Gantt bars with a today marker
                     and a progress column.
    - Milestones   : table with status chip, date, owner, description.
    - Review needed: only created if ``plan["flags"]`` is non-empty.

Design language
---------------
Editorial — no gridlines, generous row heights, restrained colour, a single
accent used sparingly (progress bars, today marker, key numbers). Palette is
derived from ``brand.primary`` and ``brand.accent`` in the plan JSON via
``brand.derive_palette``.

Usage
-----
    python generate_xlsx.py --plan plan.json --out plan.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

# Local import — scripts share a brand module for palette + logo resolution.
sys.path.insert(0, str(Path(__file__).parent))
from brand import derive_palette, resolve_logo, Palette, DEFAULT_FONT_XLSX  # noqa: E402

from openpyxl import Workbook  # noqa: E402
from openpyxl.drawing.image import Image as XLImage  # noqa: E402
from openpyxl.formatting.rule import DataBarRule  # noqa: E402
from openpyxl.styles import (  # noqa: E402
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter  # noqa: E402


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _hex(hex_with_hash: str) -> str:
    """openpyxl wants 6-digit hex without the leading ``#``."""
    return hex_with_hash.lstrip("#").upper()


def _parse_date(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def _fmt_date(d: date) -> str:
    return d.strftime("%d %b")


def _monday(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _week_columns(start: date, end: date) -> list[date]:
    """Return a list of Monday-of-week dates spanning ``start`` to ``end``."""
    cur = _monday(start)
    end_mon = _monday(end)
    weeks = []
    while cur <= end_mon:
        weeks.append(cur)
        cur += timedelta(days=7)
    return weeks


def _status_fill_colour(status: str, pal: Palette) -> str:
    return {
        "done": pal.ink_soft,
        "in_progress": pal.accent,
        "at_risk": pal.warn,
        "on_hold": pal.ink_faint,
        "not_started": pal.rule,
    }.get(status, pal.rule)


def _milestone_chip(status: str, pal: Palette) -> tuple[str, str, str]:
    return {
        "achieved":  ("Achieved",  pal.ok_soft,      pal.ok),
        "on_track":  ("On track",  pal.primary_soft, pal.primary_dark),
        "at_risk":   ("At risk",   pal.warn_soft,    pal.warn),
        "slipped":   ("Slipped",   pal.risk_soft,    pal.risk),
    }.get(status, ("Unknown", pal.hairline, pal.ink_soft))


# ---------------------------------------------------------------------------
# Sheet: Overview
# ---------------------------------------------------------------------------

def build_overview(wb: Workbook, plan: dict[str, Any], pal: Palette, logo: Path | None) -> None:
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False

    for col, w in {"A": 3, "B": 26, "C": 22, "D": 22, "E": 22, "F": 22, "G": 3}.items():
        ws.column_dimensions[col].width = w

    proj = plan["project"]
    tasks = plan.get("tasks", [])
    milestones = plan.get("milestones", [])

    total_progress = sum(float(t.get("progress", 0.0)) for t in tasks)
    pct_complete = round((total_progress / max(1, len(tasks))) * 100) if tasks else 0

    achieved = sum(1 for m in milestones if m.get("status") == "achieved")
    total_ms = len(milestones)
    at_risk_ms = sum(1 for m in milestones if m.get("status") == "at_risk")
    slipped_ms = sum(1 for m in milestones if m.get("status") == "slipped")

    if slipped_ms:
        health_label, health_text = "Off track", pal.risk
    elif at_risk_ms:
        health_label, health_text = "At risk", pal.warn
    else:
        health_label, health_text = "On track", pal.ok

    # Eyebrow
    ws.row_dimensions[2].height = 16
    ws["B2"] = proj.get("client", "").upper()
    ws["B2"].font = Font(name=DEFAULT_FONT_XLSX, size=9, bold=True, color=_hex(pal.ink_faint))

    # Title
    ws.row_dimensions[3].height = 42
    ws.merge_cells("B3:F3")
    ws["B3"] = proj["name"]
    ws["B3"].font = Font(name="Georgia", size=28, bold=True, color=_hex(pal.ink))
    ws["B3"].alignment = Alignment(vertical="center")

    # Subtitle
    ws.row_dimensions[4].height = 22
    ws.merge_cells("B4:F4")
    ws["B4"] = f"Status as of {_parse_date(proj['status_as_of']).strftime('%d %B %Y')}"
    ws["B4"].font = Font(name=DEFAULT_FONT_XLSX, size=11, italic=True, color=_hex(pal.ink_soft))

    # Hairline rule
    ws.row_dimensions[5].height = 6
    for col_letter in "BCDEF":
        ws[f"{col_letter}5"].border = Border(bottom=Side(style="thin", color=_hex(pal.rule)))

    # Snapshot cards
    ws.row_dimensions[7].height = 14
    ws.row_dimensions[8].height = 36
    ws.row_dimensions[9].height = 10

    cards = [
        ("Overall complete", f"{pct_complete}%", pal.accent),
        ("Milestones", f"{achieved} of {total_ms}", pal.primary),
        ("Schedule health", health_label, health_text),
        ("Target handover", _parse_date(proj["end"]).strftime("%d %b %Y"), pal.primary),
    ]
    for (label, value, colour), col in zip(cards, ["B", "C", "D", "E"]):
        lc = ws[f"{col}7"]
        vc = ws[f"{col}8"]
        fc = ws[f"{col}9"]
        lc.value = label.upper()
        lc.font = Font(name=DEFAULT_FONT_XLSX, size=8, bold=True, color=_hex(pal.ink_faint))
        lc.alignment = Alignment(vertical="bottom")
        vc.value = value
        vc.font = Font(name="Georgia", size=20, bold=True, color=_hex(colour))
        vc.alignment = Alignment(vertical="center")
        fc.border = Border(top=Side(style="thin", color=_hex(pal.hairline)))

    # Status summary
    ws.row_dimensions[11].height = 14
    ws["B11"] = "STATUS SUMMARY"
    ws["B11"].font = Font(name=DEFAULT_FONT_XLSX, size=8, bold=True, color=_hex(pal.ink_faint))

    ws.row_dimensions[12].height = 54
    ws.merge_cells("B12:F12")
    ws["B12"] = plan.get("status_summary", "")
    ws["B12"].font = Font(name="Georgia", size=13, italic=True, color=_hex(pal.ink))
    ws["B12"].alignment = Alignment(vertical="top", wrap_text=True)

    # Project detail
    ws.row_dimensions[14].height = 14
    ws["B14"] = "PROJECT DETAIL"
    ws["B14"].font = Font(name=DEFAULT_FONT_XLSX, size=8, bold=True, color=_hex(pal.ink_faint))

    meta_rows = [
        ("Client",          proj.get("client", "")),
        ("Project manager", proj.get("pm", "")),
        ("Start",           _parse_date(proj["start"]).strftime("%d %B %Y")),
        ("End",             _parse_date(proj["end"]).strftime("%d %B %Y")),
        ("Duration",        f"{(_parse_date(proj['end']) - _parse_date(proj['start'])).days} days"),
        ("Generated",       proj.get("generated", "")),
    ]
    for i, (k, v) in enumerate(meta_rows, start=15):
        ws.row_dimensions[i].height = 20
        ws.cell(row=i, column=2, value=k).font = Font(
            name=DEFAULT_FONT_XLSX, size=10, color=_hex(pal.ink_faint)
        )
        ws.cell(row=i, column=3, value=v).font = Font(
            name=DEFAULT_FONT_XLSX, size=11, color=_hex(pal.ink)
        )
        for col_letter in "BCDEF":
            ws[f"{col_letter}{i}"].border = Border(
                bottom=Side(style="dotted", color=_hex(pal.hairline))
            )

    # Optional logo (top right)
    if logo is not None:
        try:
            img = XLImage(str(logo))
            if img.height and img.height > 0:
                target_h = 60
                ratio = target_h / img.height
                img.height = target_h
                img.width = int(img.width * ratio)
            img.anchor = "F2"
            ws.add_image(img)
        except Exception:
            pass

    # Footer band
    footer_row = 23
    ws.row_dimensions[footer_row].height = 18
    ws.merge_cells(start_row=footer_row, start_column=2, end_row=footer_row, end_column=6)
    ws.cell(
        row=footer_row, column=2,
        value=f"Prepared by {proj.get('pm', '')} • Generated {proj.get('generated', '')}",
    ).font = Font(name=DEFAULT_FONT_XLSX, size=9, italic=True, color=_hex(pal.ink_faint))


# ---------------------------------------------------------------------------
# Sheet: Gantt
# ---------------------------------------------------------------------------

GANTT_FIRST_WEEK_COL = 6


def build_gantt(wb: Workbook, plan: dict[str, Any], pal: Palette) -> None:
    ws = wb.create_sheet("Gantt")
    ws.sheet_view.showGridLines = False

    proj = plan["project"]
    tasks = plan.get("tasks", [])
    phases = plan.get("phases", [])

    start = _parse_date(proj["start"])
    end = _parse_date(proj["end"])
    status_as_of = _parse_date(proj["status_as_of"])
    weeks = _week_columns(start, end)

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 12
    for i, _ in enumerate(weeks):
        ws.column_dimensions[get_column_letter(GANTT_FIRST_WEEK_COL + i)].width = 3.4

    # Heading
    ws.row_dimensions[2].height = 16
    ws["B2"] = proj.get("client", "").upper()
    ws["B2"].font = Font(name=DEFAULT_FONT_XLSX, size=9, bold=True, color=_hex(pal.ink_faint))

    ws.row_dimensions[3].height = 34
    ws["B3"] = "Schedule"
    ws["B3"].font = Font(name="Georgia", size=22, bold=True, color=_hex(pal.ink))

    ws.row_dimensions[4].height = 18
    ws["B4"] = f"Weekly resolution • Status as of {_parse_date(proj['status_as_of']).strftime('%d %B %Y')}"
    ws["B4"].font = Font(name=DEFAULT_FONT_XLSX, size=10, italic=True, color=_hex(pal.ink_soft))

    ws.row_dimensions[5].height = 4

    header_row_month = 6
    header_row_week = 7
    ws.row_dimensions[header_row_month].height = 18
    ws.row_dimensions[header_row_week].height = 14

    for col_letter, label in zip("BCDE", ["Task", "Owner", "Start", "Progress"]):
        c = ws[f"{col_letter}{header_row_week}"]
        c.value = label.upper()
        c.font = Font(name=DEFAULT_FONT_XLSX, size=8, bold=True, color=_hex(pal.ink_faint))
        c.alignment = Alignment(vertical="bottom")
        c.border = Border(bottom=Side(style="thin", color=_hex(pal.rule)))

    # Month banners
    current_month = None
    month_start_col = GANTT_FIRST_WEEK_COL
    for i, wk in enumerate(weeks):
        col = GANTT_FIRST_WEEK_COL + i
        if current_month is None:
            current_month = (wk.year, wk.month)
            month_start_col = col
        elif (wk.year, wk.month) != current_month:
            first_l = get_column_letter(month_start_col)
            last_l = get_column_letter(col - 1)
            if first_l != last_l:
                ws.merge_cells(f"{first_l}{header_row_month}:{last_l}{header_row_month}")
            label = date(current_month[0], current_month[1], 1).strftime("%b %Y").upper()
            mc = ws[f"{first_l}{header_row_month}"]
            mc.value = label
            mc.font = Font(name=DEFAULT_FONT_XLSX, size=9, bold=True, color=_hex(pal.ink_soft))
            mc.alignment = Alignment(horizontal="left", vertical="bottom")
            current_month = (wk.year, wk.month)
            month_start_col = col

        # week header
        wc = ws.cell(row=header_row_week, column=col, value=wk.strftime("%d"))
        wc.font = Font(name=DEFAULT_FONT_XLSX, size=7, color=_hex(pal.ink_mute))
        wc.alignment = Alignment(horizontal="center", vertical="bottom")
        wc.border = Border(bottom=Side(style="thin", color=_hex(pal.rule)))

    # final month close
    first_l = get_column_letter(month_start_col)
    last_l = get_column_letter(GANTT_FIRST_WEEK_COL + len(weeks) - 1)
    if first_l != last_l:
        ws.merge_cells(f"{first_l}{header_row_month}:{last_l}{header_row_month}")
    label = date(current_month[0], current_month[1], 1).strftime("%b %Y").upper()
    mc = ws[f"{first_l}{header_row_month}"]
    mc.value = label
    mc.font = Font(name=DEFAULT_FONT_XLSX, size=9, bold=True, color=_hex(pal.ink_soft))
    mc.alignment = Alignment(horizontal="left", vertical="bottom")

    # Today column
    today_monday = _monday(status_as_of)
    today_col: int | None = None
    for i, wk in enumerate(weeks):
        if wk == today_monday:
            today_col = GANTT_FIRST_WEEK_COL + i
            break

    # Data rows
    first_data_row = header_row_week + 1
    row = first_data_row
    progress_ranges: list[str] = []

    for phase in phases:
        phase_id = phase["id"]
        phase_name = phase["name"]
        p_start = _parse_date(phase["start"])
        p_end = _parse_date(phase["end"])

        ws.row_dimensions[row].height = 22
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        pc = ws.cell(row=row, column=2, value=phase_name.upper())
        pc.font = Font(name=DEFAULT_FONT_XLSX, size=9, bold=True, color=_hex(pal.primary_dark))
        pc.alignment = Alignment(vertical="center")

        p_start_mon = _monday(p_start)
        p_end_mon = _monday(p_end)
        for i, wk in enumerate(weeks):
            col = GANTT_FIRST_WEEK_COL + i
            cell = ws.cell(row=row, column=col)
            if p_start_mon <= wk <= p_end_mon:
                cell.fill = PatternFill("solid", fgColor=_hex(pal.primary_soft))

        for col_letter in "BCDE":
            ws[f"{col_letter}{row}"].border = Border(
                top=Side(style="thin", color=_hex(pal.rule)),
                bottom=Side(style="thin", color=_hex(pal.hairline)),
            )
        row += 1

        phase_tasks = [t for t in tasks if t.get("phase_id") == phase_id]
        phase_task_start_row = row
        for task in phase_tasks:
            ws.row_dimensions[row].height = 20

            t_start = _parse_date(task["start"])
            t_end = _parse_date(task["end"])
            progress = float(task.get("progress", 0.0))
            status = task.get("status", "not_started")

            ws.cell(row=row, column=2, value="  " + task["name"]).font = Font(
                name=DEFAULT_FONT_XLSX, size=11, color=_hex(pal.ink)
            )
            ws.cell(row=row, column=3, value=task.get("owner_role", "")).font = Font(
                name=DEFAULT_FONT_XLSX, size=10, color=_hex(pal.ink_soft)
            )
            ws.cell(row=row, column=4, value=_fmt_date(t_start)).font = Font(
                name=DEFAULT_FONT_XLSX, size=10, color=_hex(pal.ink_soft)
            )
            pcell = ws.cell(row=row, column=5, value=progress)
            pcell.number_format = "0%"
            pcell.font = Font(name=DEFAULT_FONT_XLSX, size=10, color=_hex(pal.ink))
            pcell.alignment = Alignment(horizontal="right")

            t_start_mon = _monday(t_start)
            t_end_mon = _monday(t_end)
            bar_weeks_in = [w for w in weeks if t_start_mon <= w <= t_end_mon]
            done_count = int(round(len(bar_weeks_in) * progress)) if progress > 0 else 0
            done_colour = pal.ink_soft if status == "done" else pal.accent
            rest_colour = pal.warn_soft if status == "at_risk" else pal.primary_soft

            for i, wk in enumerate(weeks):
                col = GANTT_FIRST_WEEK_COL + i
                cell = ws.cell(row=row, column=col)
                if t_start_mon <= wk <= t_end_mon:
                    idx_in_bar = bar_weeks_in.index(wk)
                    if status == "done":
                        fill = pal.ink_soft
                    elif idx_in_bar < done_count:
                        fill = done_colour
                    else:
                        fill = rest_colour
                    cell.fill = PatternFill("solid", fgColor=_hex(fill))

            for col_letter in "BCDE":
                ws[f"{col_letter}{row}"].border = Border(
                    bottom=Side(style="dotted", color=_hex(pal.hairline))
                )
            row += 1

        if row > phase_task_start_row:
            progress_ranges.append(f"E{phase_task_start_row}:E{row - 1}")

    # Today marker
    if today_col is not None:
        # Skip the merged month header row — draw the left border from the week
        # header row (header_row_week) down to the last data row.
        for r in range(header_row_week, row):
            cell = ws.cell(row=r, column=today_col)
            # Defensive: skip any merged cells we might land on
            if cell.coordinate in ws.merged_cells:
                continue
            existing = cell.border
            cell.border = Border(
                left=Side(style="medium", color=_hex(pal.accent)),
                top=existing.top,
                bottom=existing.bottom,
                right=existing.right,
            )
        # Label goes into the week header row, which is never merged.
        tlab = ws.cell(row=header_row_week, column=today_col, value="TODAY")
        tlab.font = Font(name=DEFAULT_FONT_XLSX, size=7, bold=True, color=_hex(pal.accent))
        tlab.alignment = Alignment(horizontal="left", vertical="bottom")

    # Progress data bar on each phase's task rows
    for rng in progress_ranges:
        ws.conditional_formatting.add(
            rng,
            DataBarRule(
                start_type="num", start_value=0,
                end_type="num", end_value=1,
                color=_hex(pal.accent), showValue=True,
            ),
        )

    ws.freeze_panes = f"F{first_data_row}"


# ---------------------------------------------------------------------------
# Sheet: Milestones
# ---------------------------------------------------------------------------

def build_milestones(wb: Workbook, plan: dict[str, Any], pal: Palette) -> None:
    ws = wb.create_sheet("Milestones")
    ws.sheet_view.showGridLines = False
    proj = plan["project"]
    milestones = plan.get("milestones", [])

    for col, w in {"A": 3, "B": 8, "C": 36, "D": 16, "E": 14, "F": 22, "G": 55}.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[2].height = 16
    ws["B2"] = proj.get("client", "").upper()
    ws["B2"].font = Font(name=DEFAULT_FONT_XLSX, size=9, bold=True, color=_hex(pal.ink_faint))

    ws.row_dimensions[3].height = 34
    ws["B3"] = "Milestone tracker"
    ws["B3"].font = Font(name="Georgia", size=22, bold=True, color=_hex(pal.ink))

    ws.row_dimensions[4].height = 18
    ws["B4"] = f"Status as of {_parse_date(proj['status_as_of']).strftime('%d %B %Y')}"
    ws["B4"].font = Font(name=DEFAULT_FONT_XLSX, size=10, italic=True, color=_hex(pal.ink_soft))

    header_row = 6
    ws.row_dimensions[header_row].height = 22
    for col_letter, label in zip("BCDEFG", ["ID", "Milestone", "Date", "Status", "Owner", "Description"]):
        c = ws[f"{col_letter}{header_row}"]
        c.value = label.upper()
        c.font = Font(name=DEFAULT_FONT_XLSX, size=8, bold=True, color=_hex(pal.ink_faint))
        c.alignment = Alignment(vertical="bottom")
        c.border = Border(bottom=Side(style="thin", color=_hex(pal.rule)))

    row = header_row + 1
    for m in milestones:
        ws.row_dimensions[row].height = 32
        label, fill_hex, text_hex = _milestone_chip(m.get("status", "on_track"), pal)

        ws.cell(row=row, column=2, value=m["id"]).font = Font(
            name=DEFAULT_FONT_XLSX, size=10, bold=True, color=_hex(pal.ink_soft)
        )
        ws.cell(row=row, column=3, value=m["name"]).font = Font(
            name="Georgia", size=13, color=_hex(pal.ink)
        )
        ws.cell(row=row, column=4, value=_parse_date(m["date"]).strftime("%d %b %Y")).font = Font(
            name=DEFAULT_FONT_XLSX, size=10, color=_hex(pal.ink)
        )
        status_cell = ws.cell(row=row, column=5, value=label)
        status_cell.font = Font(name=DEFAULT_FONT_XLSX, size=10, bold=True, color=_hex(text_hex))
        status_cell.fill = PatternFill("solid", fgColor=_hex(fill_hex))
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=row, column=6, value=m.get("owner_role", "")).font = Font(
            name=DEFAULT_FONT_XLSX, size=10, color=_hex(pal.ink_soft)
        )
        desc_cell = ws.cell(row=row, column=7, value=m.get("description", ""))
        desc_cell.font = Font(name=DEFAULT_FONT_XLSX, size=10, color=_hex(pal.ink_soft))
        desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

        for col_letter in "BCDEFG":
            existing = ws[f"{col_letter}{row}"].border
            ws[f"{col_letter}{row}"].border = Border(
                bottom=Side(style="dotted", color=_hex(pal.hairline)),
                left=existing.left, right=existing.right, top=existing.top,
            )
        row += 1


# ---------------------------------------------------------------------------
# Sheet: Review needed
# ---------------------------------------------------------------------------

def build_review_needed(wb: Workbook, plan: dict[str, Any], pal: Palette) -> None:
    flags = plan.get("flags", [])
    if not flags:
        return
    ws = wb.create_sheet("Review needed")
    ws.sheet_view.showGridLines = False
    proj = plan["project"]

    for col, w in {"A": 3, "B": 14, "C": 12, "D": 80}.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[2].height = 16
    ws["B2"] = proj.get("client", "").upper()
    ws["B2"].font = Font(name=DEFAULT_FONT_XLSX, size=9, bold=True, color=_hex(pal.ink_faint))

    ws.row_dimensions[3].height = 34
    ws["B3"] = "Review needed"
    ws["B3"].font = Font(name="Georgia", size=22, bold=True, color=_hex(pal.ink))

    ws.row_dimensions[4].height = 18
    ws["B4"] = "Items surfaced for explicit review before sending to the client."
    ws["B4"].font = Font(name=DEFAULT_FONT_XLSX, size=10, italic=True, color=_hex(pal.ink_soft))

    header_row = 6
    ws.row_dimensions[header_row].height = 22
    for col_letter, label in zip("BCD", ["Scope", "Ref", "Reason"]):
        c = ws[f"{col_letter}{header_row}"]
        c.value = label.upper()
        c.font = Font(name=DEFAULT_FONT_XLSX, size=8, bold=True, color=_hex(pal.ink_faint))
        c.border = Border(bottom=Side(style="thin", color=_hex(pal.rule)))

    row = header_row + 1
    for f in flags:
        ws.row_dimensions[row].height = 54
        ws.cell(row=row, column=2, value=f.get("scope", "")).font = Font(
            name=DEFAULT_FONT_XLSX, size=10, bold=True, color=_hex(pal.warn)
        )
        ws.cell(row=row, column=3, value=f.get("ref", "")).font = Font(
            name=DEFAULT_FONT_XLSX, size=10, color=_hex(pal.ink_soft)
        )
        reason_cell = ws.cell(row=row, column=4, value=f.get("reason", ""))
        reason_cell.font = Font(name=DEFAULT_FONT_XLSX, size=10, color=_hex(pal.ink))
        reason_cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col_letter in "BCD":
            ws[f"{col_letter}{row}"].border = Border(
                bottom=Side(style="dotted", color=_hex(pal.hairline))
            )
        row += 1


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Generate the project-plan Excel workbook.")
    p.add_argument("--plan", required=True, help="Path to the plan JSON file")
    p.add_argument("--out", required=True, help="Path to the output .xlsx file")
    args = p.parse_args(argv)

    plan_path = Path(args.plan)
    out_path = Path(args.out)
    with open(plan_path, "r") as f:
        plan = json.load(f)

    brand = plan.get("brand", {})
    pal = derive_palette(
        primary=brand.get("primary"),
        accent=brand.get("accent"),
        display_font=brand.get("display_font"),
        body_font=brand.get("body_font"),
    )
    logo = resolve_logo(plan.get("project", {}).get("logo_path"), plan_path)

    wb = Workbook()
    build_overview(wb, plan, pal, logo)
    build_gantt(wb, plan, pal)
    build_milestones(wb, plan, pal)
    build_review_needed(wb, plan, pal)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
